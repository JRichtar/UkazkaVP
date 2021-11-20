import pandas as pd
from slugify import slugify
from collections import defaultdict
from openpyxl import load_workbook
import math
import json

# Load the file
FILENAME = "DOD_učebny-program.xlsx"

df = pd.read_excel(FILENAME)

# Templates for badly designed .html files
TEMPLATE = """
<div class="overlay-body">
	<a href="#"><img src="assets/cross.png" id="cross" /></a>
	{0}

	{1}
</div>
"""

QUESTION = """
<div class="question {s}">
	<p>{q}</p>
	<p class="buttons">
		<a class="button answer{a_c}" href="#">{a}</a>
		<a class="button answer{b_c}" href="#">{b}</a>
		<a class="button answer{c_c}" href="#">{c}</a>
		<a class="button answer{d_c}" href="#">{d}</a>
	</p>
</div>
"""


sheet = load_workbook(FILENAME).active


# Iterate over all rows and save them to these structures
questions = defaultdict(list)
inittexts = {}
realnames = {}
unlock_points = defaultdict(int)
rooms = []

for i, row in df.iterrows():
	if not isinstance(row[1], str):
		continue

	name = slugify(row[1])
	rooms.append(name)

	realnames[name] = row[1]

	if isinstance(row["průvodní text"], str):
		inittexts[name] = row["průvodní text"]

	try:
		unlock_points[name] = int(row["počet bodů nutných ke zobrazení"])
	except ValueError:
		pass

	if row["otázka"] and row[7] and isinstance(row[7], str) and isinstance(row["otázka"], str):
		if "4" in row[7] and not "8" in row[7]:
			cls = "fouryear"
		elif "8" in row[7] and not "4" in row[7]:
			cls = "eightyear"
		else:
			cls = "both"

		if isinstance(row["Média (před otázkou)"], str) and row["Média (před otázkou)"].startswith("YT + "):
			row["Média (před otázkou)"] = row["Média (před otázkou)"][5:].lstrip()
		elif isinstance(row["Média (před otázkou)"], str) and row["Média (před otázkou)"].startswith("YT"):
			row["Média (před otázkou)"] = None
		elif isinstance(row["Média (před otázkou)"], str):
			row["Média (před otázkou)"] = row["Média (před otázkou)"].lstrip()

		if isinstance(row["Média (před otázkou)"], str) and row["Média (před otázkou)"].startswith("https://youtu.be/"):
			vidid = row["Média (před otázkou)"][17:]
			url = f"https://www.youtube.com/embed/{vidid}"

			row["Média (před otázkou)"] = f'<iframe src="{url}" frameborder="0" allow="encrypted-media" allowfullscreen></iframe>'
		elif isinstance(row["Média (před otázkou)"], str) and row["Média (před otázkou)"].startswith("http"):
			row["Média (před otázkou)"] = row["Média (před otázkou)"].split("/")[-1]
			row["Média (před otázkou)"] = '<img src="assets/questions/' + row["Média (před otázkou)"] + '" />' 
		elif isinstance(row["Média (před otázkou)"], str):
			row["Média (před otázkou)"] = '<img src="assets/questions/' + row["Média (před otázkou)"] + '" />' 
		else:
			row["Média (před otázkou)"] = ""


		q = QUESTION.format(
			s=cls, 
			q=row["Média (před otázkou)"] + row["otázka"],
			a=row["odpověď a)"], 
			b=row["odpověď b)"], 
			c=row["odpověď c)"], 
			d=row["odpověď d)"],
			a_c=" correct" if sheet.cell(i+2, 10).fill.patternType else "",
			b_c=" correct" if sheet.cell(i+2, 11).fill.patternType else "",
			c_c=" correct" if sheet.cell(i+2, 12).fill.patternType else "",
			d_c=" correct" if sheet.cell(i+2, 13).fill.patternType else "")
		questions[name].append(q)

roomsjson = {}

# Iterate over all rooms and export them to these badly designed .html files
for room in rooms:
	with open(f"assets/rooms/{room}.html", "w") as f:
		f.write(TEMPLATE.format("<em>" + inittexts.get(room, "") + "</em>", "\n".join(questions.get(room, []))))

	roomsjson[realnames[room]] = {
		"url": f"assets/rooms/{room}.html",
		"bg": f"assets/photos/{room}.jpg",
		"unlocks": unlock_points[room],
		"tooltip": realnames[room] + (
			(" &mdash; " + inittexts[room]) if inittexts.get(room) else ""
		) + (
			f" &mdash; Je potřeba {unlock_points[room]} bodů pro vstup." if unlock_points[room] > 0 else ""
		)
	}

with open("assets/rooms.json", "w") as f:
	json.dump(roomsjson, f, indent=1, sort_keys=True)
